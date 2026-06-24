import argparse
import json
import re
import uuid
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = ROOT.parent / "THU MUC LINK COPY"

ACCOUNT_FILE = "Copy of 1. [HPUS-KT210] USBC101- ACCOUNT BALANCE 2026.xlsx"
PC49_RC_FILE = "Copy of 5. [HPUS-KT210] US_PC49 RC JM (FINAL) 2026.xlsx"
TRANS_RC_FILE = "Copy of 6. [HPUS-KT210] TRANS RC JM (FINAL) 2026.xlsx"

ACCOUNT_SHEETS = {
    "1.Trans": "Trans",
    "2.Pc49": "PC49",
    "3.TDW": "TDW",
    "4.HPLLC": "HPLLC",
    "5.3NVY": "3NVY",
    "6.Other": "Other",
}

NAMESPACE = uuid.UUID("5cda246f-b569-4d9b-9385-5160a2a18c75")


def clean_text(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def lower_text(value):
    return clean_text(value).lower() if clean_text(value) else ""


def to_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, str):
        if value.strip().startswith("#"):
            return Decimal("0")
        value = value.replace(",", "").replace("$", "").strip()
        if not value:
            return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def money(value):
    return float(to_decimal(value).quantize(Decimal("0.01")))


def has_money(value):
    return abs(money(value)) > 0


def date_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value or value.startswith("#"):
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def normalize_company(value, fallback=None):
    text = clean_text(value) or fallback or "Other"
    lookup = {
        "pc49": "PC49",
        "pc 49": "PC49",
        "trans": "Trans",
        "tfj": "Trans",
        "tdw": "TDW",
        "hpllc": "HPLLC",
        "3nvy": "3NVY",
        "other": "Other",
    }
    return lookup.get(text.lower(), text)


def normalize_type(value, desc=None, expense=0, receipt=0, deposit=0, rc_no=None):
    raw = lower_text(value)
    desc_raw = lower_text(desc)
    if "extra" in raw and "deposit" in raw:
        return "extra_deposit"
    if "pick" in raw:
        return "pick_up"
    if "deposit" in raw:
        return "deposit"
    if "receipt" in raw:
        return "receipt"
    if raw in {"po", "purchase"} or "mua vao" in desc_raw or "mua vào" in desc_raw:
        return "po"
    if "return" in raw:
        return "return"
    if "exchange" in raw or "doi" in desc_raw or "đổi" in desc_raw:
        return "exchange"
    if "transfer" in raw:
        return "transfer"
    if "repair" in raw:
        return "repair"
    if expense:
        return "po"
    rc = clean_text(rc_no) or ""
    if deposit or rc.startswith("900"):
        return "deposit"
    if "pickup" in desc_raw or "pick up" in desc_raw:
        return "pick_up"
    return "receipt"


def status_for(tx_type, desc):
    desc_raw = lower_text(desc)
    if "cancel" in desc_raw:
        return "cancel"
    if tx_type in {"deposit", "extra_deposit"}:
        return "dat_coc"
    if tx_type == "return":
        return "return"
    if tx_type == "exchange":
        return "exchange"
    return "hoan_tat"


def clean_source(value):
    text = clean_text(value)
    if not text:
        return None
    if text.lower() in {"khong co source", "không có source", "khong có source"}:
        return None
    return text


def clean_rc(value):
    text = clean_text(value)
    if not text:
        return None
    compact = re.sub(r"\D", "", text)
    if len(compact) >= 7 and (compact.startswith("100") or compact.startswith("900")):
        return compact
    return None


def parse_pct(value):
    if value is None:
        return None
    if isinstance(value, str) and "%" in value:
        number = money(value.replace("%", ""))
        return number if number > 1 else number * 100
    number = money(value)
    if number == 0:
        return None
    return number * 100 if 0 < number <= 1 else number


def key_for(row):
    return (
        row.get("company"),
        row.get("ngay"),
        lower_text(row.get("dien_giai")),
        lower_text(row.get("khach")),
        round(row.get("expense", 0) or 0, 2),
        round((row.get("ar_cash", 0) or 0) + (row.get("ar_bankwire", 0) or 0) + (row.get("ar_zelle", 0) or 0) + (row.get("ar_check", 0) or 0), 2)
        if row.get("type") in {"receipt", "pick_up", "repair"}
        else 0,
        round((row.get("ar_cash", 0) or 0) + (row.get("ar_bankwire", 0) or 0) + (row.get("ar_zelle", 0) or 0) + (row.get("ar_check", 0) or 0), 2)
        if row.get("type") in {"deposit", "extra_deposit"}
        else 0,
    )


def stable_id(source):
    return str(uuid.uuid5(NAMESPACE, source))


def add_line_and_payment(record, source_name):
    amount = 0
    if record["type"] in {"receipt", "pick_up", "repair", "deposit", "extra_deposit"}:
        amount = record["ar_cash"] + record["ar_bankwire"] + record["ar_zelle"] + record["ar_check"]
    elif record["type"] in {"po", "return", "exchange"}:
        amount = record["expense"]

    if abs(amount) > 0:
        record["line_items"] = [{
            "mo_ta": record.get("dien_giai") or source_name,
            "sku": record.get("ma_sku"),
            "gia_no": None,
            "so_luong": 1,
            "don_gia": round(amount, 2),
        }]
    else:
        record["line_items"] = []

    record["payments"] = []
    paid = record["ar_cash"] + record["ar_bankwire"] + record["ar_zelle"] + record["ar_check"]
    if record["type"] in {"receipt", "pick_up", "repair", "deposit", "extra_deposit"} and abs(paid) > 0:
        buckets = [
            ("cash", record["ar_cash"]),
            ("bank_wire", record["ar_bankwire"]),
            ("zelle", record["ar_zelle"]),
            ("check", record["ar_check"]),
        ]
        for method, value in buckets:
            if abs(value) > 0:
                record["payments"].append({
                    "ngay": record["ngay"],
                    "so_tien": round(value, 2),
                    "hinh_thuc": method,
                    "nguoi_xac_nhan": None,
                    "ghi_chu": source_name,
                    "is_dau": record["type"] in {"deposit", "extra_deposit"},
                })


def parse_account_balance():
    records = []
    wb = load_workbook(EXCEL_DIR / ACCOUNT_FILE, read_only=True, data_only=True)
    for sheet, fallback_company in ACCOUNT_SHEETS.items():
        ws = wb[sheet]
        for idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
            no = row[1] if len(row) > 1 else None
            ngay = date_iso(row[2] if len(row) > 2 else None)
            tx_type_raw = row[3] if len(row) > 3 else None
            desc = clean_text(row[4] if len(row) > 4 else None)
            company = normalize_company(row[9] if len(row) > 9 else None, fallback_company)
            if not isinstance(no, (int, float)) or not ngay or not tx_type_raw or not desc:
                continue

            condition_return_po = money(row[11] if len(row) > 11 else None)
            condition_receipt = money(row[12] if len(row) > 12 else None)
            condition_deposit = money(row[13] if len(row) > 13 else None)
            ar_cash = money(row[14] if len(row) > 14 else None)
            ar_bankwire = money(row[15] if len(row) > 15 else None)
            ar_zelle = money(row[16] if len(row) > 16 else None)
            ar_check = money(row[17] if len(row) > 17 else None)

            tx_type = normalize_type(
                tx_type_raw,
                desc,
                condition_return_po,
                condition_receipt,
                condition_deposit,
            )
            expected = 0
            if tx_type in {"receipt", "pick_up", "repair"}:
                expected = condition_receipt
            elif tx_type in {"deposit", "extra_deposit"}:
                expected = condition_deposit
            current = ar_cash + ar_bankwire + ar_zelle + ar_check
            if expected and round(current, 2) != round(expected, 2):
                ar_cash = round(ar_cash + (expected - current), 2)

            record = {
                "id": stable_id(f"account:{sheet}:{idx}"),
                "ngay": ngay,
                "company": company,
                "type": tx_type,
                "ma_sku": clean_text(row[5] if len(row) > 5 else None),
                "dien_giai": desc,
                "khach": clean_text(row[7] if len(row) > 7 else None) or "Không có thông tin",
                "contact": clean_text(row[8] if len(row) > 8 else None),
                "expense": condition_return_po if tx_type in {"po", "return", "exchange"} else 0,
                "ar_cash": ar_cash,
                "ar_bankwire": ar_bankwire,
                "ar_zelle": ar_zelle,
                "ar_check": ar_check,
                "rc_jm_no": None,
                "so_no": None,
                "appt_id": None,
                "source_1": None,
                "source_2": None,
                "sale_1": None,
                "sale_online": None,
                "transaction_value": None,
                "pct_support": None,
                "old_receipt_no": None,
                "deposit_date": None,
                "bell_code": clean_text(row[6] if len(row) > 6 else None),
                "trang_thai": status_for(tx_type, desc),
                "note": f"Import {ACCOUNT_FILE} / {sheet} row {idx}",
                "_source": f"account:{sheet}:{idx}",
            }
            add_line_and_payment(record, "Account Balance")
            records.append(record)
    return records


def rc_base_record(company, row, sheet_name, idx, layout):
    if layout == "pc49":
        no, ngay, desc, customer = row[0], row[1], row[2], row[3]
        expense, receipt, deposit, rc_no = row[4], row[5], row[6], row[7]
        source_1, source_2 = row[8], row[9]
        sale_1 = row[10]
        sale_2 = row[11]
        sale_online = row[12]
        transaction_value = row[15]
        pct_support = row[16]
        note = row[17]
        ma_sku = None
        bell_code = None
        old_receipt_no = None
        deposit_date = None
        contact = None
    else:
        no, ngay, desc = row[0], row[1], row[2]
        ma_sku, bell_code, customer = row[3], row[4], row[5]
        expense, receipt, deposit, rc_no = row[6], row[7], row[8], row[9]
        source_1, source_2 = row[10], row[11]
        sale_1 = row[13]
        sale_2 = row[16]
        sale_3 = row[19] if len(row) > 19 else None
        sale_online = "; ".join([x for x in [clean_text(row[20] if len(row) > 20 else None), clean_text(row[21] if len(row) > 21 else None), clean_text(row[22] if len(row) > 22 else None)] if x]) or None
        transaction_value = row[23] if len(row) > 23 else None
        pct_support = row[24] if len(row) > 24 else None
        old_receipt_no = row[25] if len(row) > 25 else None
        deposit_date = row[26] if len(row) > 26 else None
        contact = row[28] if len(row) > 28 else None
        note = "; ".join([x for x in [clean_text(row[30] if len(row) > 30 else None), clean_text(row[31] if len(row) > 31 else None), clean_text(row[32] if len(row) > 32 else None)] if x]) or None
        sale_2 = "; ".join([x for x in [clean_text(sale_2), clean_text(sale_3)] if x]) or None

    if not isinstance(no, (int, float)) or not date_iso(ngay) or not clean_text(desc):
        return None

    expense_value = money(expense)
    receipt_value = money(receipt)
    deposit_value = money(deposit)
    tx_type = normalize_type(None, desc, expense_value, receipt_value, deposit_value, rc_no)
    rc_clean = clean_rc(rc_no)
    raw_rc = clean_text(rc_no)

    record = {
        "id": stable_id(f"rc:{layout}:{sheet_name}:{idx}"),
        "ngay": date_iso(ngay),
        "company": company,
        "type": tx_type,
        "ma_sku": clean_text(ma_sku),
        "dien_giai": clean_text(desc),
        "khach": clean_text(customer) or "Không có thông tin",
        "contact": clean_text(contact),
        "expense": expense_value if tx_type in {"po", "return", "exchange"} else 0,
        "ar_cash": receipt_value + deposit_value if tx_type in {"receipt", "pick_up", "repair", "deposit", "extra_deposit"} else 0,
        "ar_bankwire": 0,
        "ar_zelle": 0,
        "ar_check": 0,
        "rc_jm_no": rc_clean,
        "so_no": None,
        "appt_id": None,
        "source_1": clean_source(source_1),
        "source_2": clean_source(source_2),
        "sale_1": "; ".join([x for x in [clean_text(sale_1), clean_text(sale_2)] if x]) or None,
        "sale_online": clean_text(sale_online),
        "transaction_value": clean_text(transaction_value),
        "pct_support": parse_pct(pct_support),
        "old_receipt_no": clean_text(old_receipt_no),
        "deposit_date": date_iso(deposit_date),
        "bell_code": clean_text(bell_code),
        "trang_thai": status_for(tx_type, desc),
        "note": "; ".join([x for x in [clean_text(note), f"Import {sheet_name} row {idx}", f"Raw receipt: {raw_rc}" if raw_rc and not rc_clean else None] if x]),
        "_source": f"rc:{layout}:{sheet_name}:{idx}",
    }
    add_line_and_payment(record, f"RC JM {company}")
    return record


def parse_rc_files():
    records = []
    configs = [
        (PC49_RC_FILE, "PC49", "pc49", "Receipt PC49-", 4),
        (TRANS_RC_FILE, "Trans", "trans", "Receipt TFJ-", 7),
    ]
    for file_name, company, layout, prefix, start_row in configs:
        wb = load_workbook(EXCEL_DIR / file_name, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith(prefix):
                continue
            ws = wb[sheet_name]
            for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                record = rc_base_record(company, row, sheet_name, idx, layout)
                if record:
                    records.append(record)
    return records


def merge_records(account_records, rc_records):
    account_pool = defaultdict(list)
    for record in account_records:
        account_pool[key_for(record)].append(record)

    merged = []
    matched_sources = set()
    used_rc = set()

    for rc in rc_records:
        match = None
        pool = account_pool.get(key_for(rc))
        if pool:
            match = pool.pop(0)
            matched_sources.add(match["_source"])

        if match:
            rc["id"] = match["id"]
            rc["contact"] = rc.get("contact") or match.get("contact")
            rc["ma_sku"] = rc.get("ma_sku") or match.get("ma_sku")
            rc["bell_code"] = rc.get("bell_code") or match.get("bell_code")
            rc["ar_cash"] = match["ar_cash"]
            rc["ar_bankwire"] = match["ar_bankwire"]
            rc["ar_zelle"] = match["ar_zelle"]
            rc["ar_check"] = match["ar_check"]
            rc["expense"] = match["expense"]
            rc["type"] = match["type"]
            rc["trang_thai"] = status_for(rc["type"], rc["dien_giai"])
            rc["note"] = "; ".join([x for x in [rc.get("note"), match.get("note")] if x])
            add_line_and_payment(rc, f"RC JM {rc['company']} + Account Balance")

        rc_no = rc.get("rc_jm_no")
        if rc_no:
            if rc_no in used_rc:
                rc["note"] = "; ".join([x for x in [rc.get("note"), f"Duplicate RC/JM ignored: {rc_no}"] if x])
                rc["rc_jm_no"] = None
            else:
                used_rc.add(rc_no)
        merged.append(rc)

    for account in account_records:
        if account["_source"] in matched_sources:
            continue
        merged.append(account)

    for record in merged:
        record.pop("_source", None)
    return merged


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(ROOT / "scripts" / "import-output.json"))
    args = parser.parse_args()

    account_records = parse_account_balance()
    rc_records = parse_rc_files()
    merged = merge_records(account_records, rc_records)

    summary = {
        "account_balance_rows": len(account_records),
        "rc_jm_rows": len(rc_records),
        "transactions": len(merged),
        "payments": sum(len(r["payments"]) for r in merged),
        "line_items": sum(len(r["line_items"]) for r in merged),
        "by_company": {},
        "by_type": {},
    }
    for record in merged:
        summary["by_company"][record["company"]] = summary["by_company"].get(record["company"], 0) + 1
        summary["by_type"][record["type"]] = summary["by_type"].get(record["type"], 0) + 1

    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "sources": [ACCOUNT_FILE, PC49_RC_FILE, TRANS_RC_FILE],
        "summary": summary,
        "transactions": merged,
    }
    out = Path(args.out)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
