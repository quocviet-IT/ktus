import argparse
import json
import uuid
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXCEL_DIR = ROOT.parent / "THU MUC LINK COPY"
NAMESPACE = uuid.UUID("c7ef28ec-c1ee-4077-9401-91579f3436dd")


def stable_id(kind, *parts):
    return str(uuid.uuid5(NAMESPACE, ":".join([kind, *map(str, parts)])))


def serialize_cell(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).strip()


def row_values(row):
    values = [serialize_cell(value) for value in row]
    while values and (values[-1] is None or values[-1] == ""):
        values.pop()
    return values


def row_text(values):
    return " | ".join(str(value) for value in values if value not in (None, ""))


def write_record(handle, record):
    handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=str(ROOT / "scripts" / "raw-excel-output.jsonl"))
    args = parser.parse_args()

    out = Path(args.out)
    files = sorted(EXCEL_DIR.glob("*.xlsx"))
    totals = {"workbooks": 0, "sheets": 0, "rows": 0, "failures": []}

    with out.open("w", encoding="utf-8") as handle:
        for file_path in files:
            workbook_id = stable_id("workbook", file_path.name)
            stat = file_path.stat()
            try:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                write_record(handle, {
                    "kind": "workbook",
                    "id": workbook_id,
                    "file_name": file_path.name,
                    "relative_path": str(file_path.relative_to(EXCEL_DIR.parent)),
                    "file_size_bytes": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                })
                totals["workbooks"] += 1

                for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
                    sheet_id = stable_id("sheet", file_path.name, sheet.title)
                    write_record(handle, {
                        "kind": "sheet",
                        "id": sheet_id,
                        "workbook_id": workbook_id,
                        "sheet_name": sheet.title,
                        "sheet_index": sheet_index,
                        "max_row": sheet.max_row or 0,
                        "max_column": sheet.max_column or 0,
                    })
                    totals["sheets"] += 1

                    non_empty = 0
                    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        values = row_values(row)
                        if not values or not any(value not in (None, "") for value in values):
                            continue
                        non_empty += 1
                        write_record(handle, {
                            "kind": "row",
                            "id": stable_id("row", file_path.name, sheet.title, row_index),
                            "workbook_id": workbook_id,
                            "sheet_id": sheet_id,
                            "row_index": row_index,
                            "cells": values,
                            "row_text": row_text(values),
                        })
                    totals["rows"] += non_empty
                    print(f"{file_path.name} / {sheet.title}: {non_empty} rows")
            except Exception as exc:
                totals["failures"].append({"file_name": file_path.name, "error": str(exc)})

    print(json.dumps(totals, ensure_ascii=False, indent=2))
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
